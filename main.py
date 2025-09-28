from threading import Thread
from openpyxl import load_workbook
from pygame import Vector2
from PIL import Image
import pygame, sys, os, datetime
import xlrd
import pygame.freetype as ft
import yaml

pygame.init()
ft.init()


def get_png_size(image_path):
    """
    Opens a PNG image and returns its width and height.
    Returns (width, height) tuple or None if an error occurs.
    """
    try:
        # 1. Open the image file
        with Image.open(image_path) as img:
            # 2. Access the .size attribute
            width, height = img.size
            return width, height
    except FileNotFoundError:
        print(f"Error: The file '{image_path}' was not found.")
        return None
    except Exception as e:
        print(f"An error occurred: {e}")
        return None


PATH = os.path.dirname(os.path.abspath(__file__))
config = yaml.full_load(open(os.path.join(PATH, "export-config.yml"), "r"))

default_img_size = get_png_size(os.path.join(PATH, "assets", "default.png"))
if default_img_size:
    WIDTH, HEIGHT = default_img_size
else:
    print("Failed to get default image size, exiting...")
    exit(1)
window = pygame.display.set_mode((WIDTH, HEIGHT))
running = True
clock = pygame.time.Clock()
DEBUGGING = config.get("DEBUG", False)

BG_COLOR = (255, 255, 190)  # also for school name color
INFO_COLOR = (23, 121, 254)


start_row: int = config.get("start_row", 4)
name_col: int = config.get("name_col", 1)
split_name: bool = config.get("split_name", False)
gender_col: int = config.get("gender_col", 4)
dob_col: int = config.get("dob_col", 4)
class_name: str = config.get("class_name", "1A")
file_name: str = config.get("file_name", "data")

"""//////////////CONST POSITION FOR GENERATION (DIFFERENT FOR EACH DEFAULT IMAGE)//////////////////"""
NAME_POS = Vector2(config.get('NAME_POS_X'), config.get('NAME_POS_Y'))
DOB_POS = Vector2(config.get('DOB_POS_X'), config.get('DOB_POS_Y'))
NK_POS = Vector2(config.get('NK_POS_X'), config.get('NK_POS_Y'))
SCHOOL_POS = Vector2(50, 50)
GENDER_POS = ()
FONT_SIZE = 47
"""//////////////CONST POSITION FOR GENERATION (DIFFERENT FOR EACH DEFAULT IMAGE)//////////////////"""

"""//////////////CURRENT CAPTURING INFO//////////////////"""
cur_row_id = start_row
student_index = 1
cur_name = ""
cur_dob = ""
cur_gender = ""
"""//////////////CURRENT CAPTURING INFO//////////////////"""

# Load the workbook
try:  # xlsx first
    workbook = load_workbook(
        filename=os.path.join(PATH, "excel_files", f"{file_name}.xlsx"), read_only=True
    )
    if len(workbook.sheetnames) > 1:
        print(
            f"WARNING: There's more than 1 sheet in {file_name}, manual intervention if needed"
        )
    sheet = workbook.active
    if sheet == None:
        raise Exception("No active sheet found")
    data_list = [
        [
            (
                cell.strftime("%d/%m/%Y")
                if isinstance(cell, datetime.datetime)
                else str(cell)
            )
            for cell in row
        ]
        for row in sheet.iter_rows(values_only=True)
    ]
except Exception as e:  # try xls if xlsx failed

    print(f"Failed to load .xlsx file due to {e}, trying .xls file...")

    workbook = xlrd.open_workbook(os.path.join(PATH, "excel_files", f"{file_name}.xls"))
    if len(workbook.sheet_names()) > 1:
        print(
            f"WARNING: There's more than 1 sheet in {file_name}, manual intervention if needed"
        )
    if len(workbook.sheet_names()) == 0:
        raise Exception("No sheet found")

    sheet = workbook.sheet_by_index(0)
    data_list = [
        [(sheet.cell_value(r, c)) for c in range(sheet.ncols)]
        for r in range(sheet.nrows)
    ]


"""//////////////CONST////////////////"""
INFO_FONT = ft.Font(os.path.join(PATH, "assets", "TNR.ttf"), FONT_SIZE)
default_img = pygame.image.load(os.path.join(PATH, "assets", "default.png")).convert()
INFO_STYLE = ft.STYLE_STRONG

cur_nk = config.get("school_year", "2023 - 2028")


def update():  # to load the current student data
    pygame.display.set_caption(str(int(clock.get_fps())))
    global cur_name_surf, cur_dob_surf, cur_name, cur_dob
    try:
        cur_name = (
            data_list[cur_row_id][name_col]
            + split_name
            * (
                " " * (data_list[cur_row_id][name_col][-1] != " ")
                + data_list[cur_row_id][name_col + 1]
            )
        ).title()
    except:
        exit(0)
    cur_dob = data_list[cur_row_id][dob_col]
    # cur_name_surf = INFO_FONT.render(
    #     cur_name, INFO_COLOR, style=INFO_STYLE, size=FONT_SIZE
    # )[0]
    # cur_dob_surf = INFO_FONT.render(
    #     cur_dob, INFO_COLOR, style=INFO_STYLE, size=FONT_SIZE
    # )[0]


def calculate_corrected_ascender(font_obj, text, size):
    """
    Calculates the actual highest point (ascender) needed for a text string
    by checking the metrics of all individual glyphs.
    """
    # get_metrics returns a list of tuples:
    # [(min_x, max_x, min_y, max_y, advance), ...]
    metrics = font_obj.get_metrics(text, size=size)

    # max_y_offset is the fourth element (index 3). We find the maximum
    # max_y_offset across all characters in the string.
    max_y = 0
    for glyph_metrics in metrics:
        # glyph_metrics can be None if the character isn't supported
        if glyph_metrics is not None:
            max_y = max(max_y, glyph_metrics[3])

    return max_y


def blit_corrected_baseline(
    surf: pygame.Surface, font: ft.Font, text: str, dst: Vector2
):
    corrected_ascender = calculate_corrected_ascender(font, text, FONT_SIZE)

    BASE_Y_CORRECTED = dst.y

    blit_y_corrected = BASE_Y_CORRECTED - corrected_ascender

    # 4. Render and Blit the text
    text_surf_corrected = font.render(
        text, INFO_COLOR, size=FONT_SIZE, style=INFO_STYLE
    )[0]
    surf.blit(text_surf_corrected, (dst.x, blit_y_corrected))


def draw():  # to draw the data on the screen
    window.blit(default_img, (0, 0))
    # window.blit(cur_name_surf, NAME_POS - Vector2(0, cur_name_surf.get_height() / 2))
    # window.blit(cur_dob_surf, DOB_POS - Vector2(0, cur_dob_surf.get_height() / 2))
    # window.blit(cur_nk_surf, NK_POS - Vector2(0, cur_nk_surf.get_height() / 2))

    blit_corrected_baseline(window, INFO_FONT, cur_name, NAME_POS)
    blit_corrected_baseline(window, INFO_FONT, cur_dob, DOB_POS)
    blit_corrected_baseline(window, INFO_FONT, cur_nk, NK_POS)

    # window.blit(cur_school_surf,SCHOOL_POS)


def capture():  # to save the image
    global student_index, cur_row_id
    if not os.path.exists(os.path.join(PATH, "result", class_name)):
        os.mkdir(
            os.path.join(PATH, "result", class_name),
        )
    # pygame.image.save(
    #     window, os.path.join(PATH, "result", class_name, f"{cur_name}.png")
    # )
    if cur_name == None or cur_name == "None":
        global running
        running = False
        return
    t = Thread(
        target=pygame.image.save,
        args=(
            window,
            os.path.join(PATH, "result", class_name, f"{cur_name}.png"),
        ),
    )
    t.start()
    if DEBUGGING:
        t.join()
        # pygame.time.wait(10000)
        exit(0)
    else:
        student_index += 1
        cur_row_id += 1


if __name__ == "__main__":
    while running:
        for event in pygame.event.get():
            if event.type == pygame.QUIT:
                pygame.quit()
                sys.exit()
        update()
        draw()
        capture()
        pygame.display.flip()
        clock.tick(60)
