import os
from openpyxl import load_workbook

PATH = os.path.dirname(os.path.abspath(__file__)).replace("\\","/") + "/excel_files/"

# Load the workbook
workbook = load_workbook(filename = PATH + "data2.xlsx")

# Select the first sheet
sheet = workbook.active

# Get the data
data_list = [[cell for cell in row] for row in sheet.iter_rows(values_only=True)]

for row in data_list:
    print(row)
